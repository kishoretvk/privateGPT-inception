from openpyxl import load_workbook
from typing import Dict, List, Optional
from langchain.docstore.document import Document
from langchain.document_loaders.base import BaseLoader
from multiprocessing import Pool # Import multiprocessing module

class XLSXLoader(BaseLoader):
   """Loads an XLSX file into a list of documents.
   Each document represents one row of the XLSX file. Every row is converted into a
   key/value pair and outputted to a new line in the document's page_content.
   The source for each document loaded from xlsx is set to the value of the
   'file_path' argument for all documents by default.
   You can override this by setting the 'source_column' argument to the
   name of a column in the XLSX file.
   The source of each document will then be set to the value of the column
   with the name specified in 'source_column'.
   Output Example:
       .. code-block:: txt
           column1: value1
           column2: value2
           column3: value3
   """

   def __init__(
           self,
           file_path: str,
           source_column: Optional[str] = None,
           sheet_name: Optional[str] = None,
           encoding: Optional[str] = None,
   ):
      self.file_path = file_path
      self.source_column = source_column
      self.sheet_name = sheet_name
      self.encoding = encoding

   def load(self) -> List[Document]:
    docs = []

    # Load the workbook object and get the sheet names
    wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames if self.sheet_name is None else [self.sheet_name]

    # Create a pool of worker processes
    pool = Pool()

    # Define a function that takes a sheet name and returns a list of documents from that sheet
    def create_docs(sheet_name):
        # Get the worksheet object and the headers from the first row
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        # Create a list of documents from each row using the Document class
        docs = []
        for row in ws.iter_rows(min_row=2):
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(headers, row_values))

            # Update this line to fix the error and use built-in functions
            content = "\n".join(f"{k.strip()}: {v.strip() if isinstance(v, str) else v}" for k, v in row_dict.items() if v is not None)

            if self.source_column is not None:
                source = row_dict[self.source_column]
            else:
                source = self.file_path

            metadata = {"source": source}
            doc = Document(page_content=content, metadata=metadata)
            docs.append(doc)

        return docs

    # Use map to apply create_docs function to each sheet name in parallel and get a list of lists of documents
    docs_list = pool.map(create_docs, sheet_names)

    # Flatten the list of lists into a single list using itertools.chain
    from itertools import chain
    docs = list(chain.from_iterable(docs_list))

    # Close and join the pool
    pool.close()
    pool.join()

    return docs